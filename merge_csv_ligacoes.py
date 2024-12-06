import glob
import os.path
import pandas as pd
from openpyxl import load_workbook

# Caminho da pasta onde estão os arquivos CSV
folder_path = r"C:\Users\Dell\Downloads"
file_type = '/*csv'
files = glob.glob(folder_path + file_type)

# Seleciona os 4 arquivos CSV mais recentes
files.sort(key=os.path.getctime, reverse=True)  # Ordena do mais recente para o mais antigo
recent_files = files[:4]  # Pega os 4 primeiros

print("Arquivos selecionados para processamento:")
for file in recent_files:
    print(file)

# Lista para consolidar todos os DataFrames
all_dataframes = []

# Lê cada arquivo CSV e adiciona ao DataFrame consolidado
for file in recent_files:
    try:
        df = pd.read_csv(file, encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(file, encoding='latin1')
    all_dataframes.append(df)

# Junta todos os DataFrames em um único
df_concatenado = pd.concat(all_dataframes, ignore_index=True)

# Mapeamento de extensionNumber para nomes
name_mapping = {
    1001: "Dayara",
    1005: "Fabrizia",
    1006: "Thais",
    1013: "Giovanna"
}

# Consolidação dos dados
consolidado = df_concatenado.groupby(['extensionNumber', 'duration']).size().reset_index(name='contagem')

# Adiciona os nomes dos colaboradores com base no extensionNumber
consolidado['colaborador'] = consolidado['extensionNumber'].map(name_mapping)

# --- BLOCO ADICIONADO: Transpor os dados para formato de colunas ---
# Transforma os dados para o formato desejado
consolidado_pivot = consolidado.pivot_table(
    index='colaborador',   # Colunas que serão os índices (nomes dos colaboradores)
    columns='duration',    # Valores que serão transformados em colunas (duração)
    values='contagem',     # Valores a serem preenchidos (contagem das ligações)
    aggfunc='sum',         # Agregação caso existam duplicatas
    fill_value=0           # Preencher valores faltantes com 0
).reset_index()

# Ordena as colunas pelo tempo de duração, se necessário
column_order = ['colaborador'] + sorted(consolidado_pivot.columns[1:], key=lambda x: (x == 'mais que 03:01', x))
consolidado_pivot = consolidado_pivot[column_order]

# Calcula o total de ligações por colaborador
consolidado_pivot['Total Geral'] = consolidado_pivot.iloc[:, 1:].sum(axis=1)

# Salva os dados no Excel
output_file = 'Marillac - Relatórios das ligações de ontem.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
    consolidado_pivot.to_excel(writer, index=False, sheet_name='Consolidado')

# Adiciona a soma total geral na célula A7
wb = load_workbook(output_file)
sheet = wb['Consolidado']

# Calcula a soma total das chamadas e insere na célula A7
total_calls = consolidado_pivot['Total Geral'].sum()
sheet.cell(row=7, column=1, value="Total Geral de Chamadas")
sheet.cell(row=7, column=2, value=total_calls)

# Salva o arquivo com a soma total adicionada
wb.save(output_file)

print(f"Dados transpostos e total geral adicionado no arquivo {output_file}, na aba 'Consolidado'.")
